from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class Excel:

    HEADER_FISSI = [
        'Azienda', 'File sorgente',
        'Cognome', 'Nome', 'ComuneNascita', 'DataNascita', 'Sesso',
        'CF', 'ComuneResidenza', 'PR', 'CAP', 'Indirizzo',
        'DataAssunzione', 'Note', 'Mansioni', 'Email',
        'Matricola', 'Titolo', 'Nazionalita', 'ProvinciaNascita',
        'FormazioneCogente', 'SorveglianzaSanitaria'
    ]
    

    def __init__(self):
        pass

    def crea_excel(self, tutte_risorse, percorso_output):
        """Crea il file Excel con foglio RIEPILOGO + un foglio per azienda."""
        wb = Workbook()

        # Stili
        COLORE_HEADER   = "1F4E79"
        COLORE_ALT_RIGA = "DEEAF1"
        COLORE_CONS     = "FFF2CC"  # giallo chiaro per consulenti (Tipo != 1)

        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell_font   = Font(name="Arial", size=10)
        header_fill = PatternFill("solid", start_color=COLORE_HEADER)
        alt_fill    = PatternFill("solid", start_color=COLORE_ALT_RIGA)
        cons_fill   = PatternFill("solid", start_color=COLORE_CONS)
        border_thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        center = Alignment(horizontal="center", vertical="center")
        left   = Alignment(horizontal="left",   vertical="center")

        LARGHEZZE = {
            'Azienda': 28, 'File sorgente': 24, 'Cognome': 18, 'Nome': 16,
            'ComuneNascita': 18, 'DataNascita': 13, 'Sesso': 10,
            'CF': 18, 'ComuneResidenza': 18, 'PR': 6, 'CAP': 8,
            'Indirizzo': 24, 'DataAssunzione': 15, 'Note': 22,
            'Mansioni': 28, 'Email': 26, 'Matricola': 10, 'Titolo': 10,
            'Nazionalita': 14, 'ProvinciaNascita': 14,
            'FormazioneCogente': 16, 'SorveglianzaSanitaria': 18,
        }

        def scrivi_foglio(ws, righe, titolo):
            ws.title = titolo[:31]
            ws.freeze_panes = "A2"

            # Intestazioni
            for ci, h in enumerate(self.HEADER_FISSI, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center
                cell.border    = border_thin
            ws.row_dimensions[1].height = 20

            # Dati
            for ri, risorsa in enumerate(righe, 2):
                # Colore riga: giallo per non-lavoratori, alternato per lavoratori
                tipo = risorsa.get('_tipo', 1)
                if tipo != 1:
                    fill = cons_fill
                else:
                    fill = alt_fill if ri % 2 == 0 else PatternFill()

                for ci, h in enumerate(self.HEADER_FISSI, 1):
                    val  = risorsa.get(h, '')
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font      = cell_font
                    cell.fill      = fill
                    cell.alignment = left
                    cell.border    = border_thin

            # Larghezze colonne
            for ci, h in enumerate(self.HEADER_FISSI, 1):
                ws.column_dimensions[get_column_letter(ci)].width = LARGHEZZE.get(h, 14)

        # Foglio RIEPILOGO
        ws_riep = wb.active
        scrivi_foglio(ws_riep, tutte_risorse, "RIEPILOGO")

        # Un foglio per azienda
        aziende = {}
        for r in tutte_risorse:
            az = r.get('Azienda', 'Sconosciuta')
            aziende.setdefault(az, []).append(r)

        for nome_az, righe in sorted(aziende.items()):
            nome_pulito = "".join(
                c for c in nome_az if c not in r'\/*?:[]\x00-\x1f'
            )[:31] or "Azienda"
            ws = wb.create_sheet(title=nome_pulito)
            scrivi_foglio(ws, righe, nome_pulito)

        wb.save(percorso_output)
        print(f"  File salvato: {percorso_output}")