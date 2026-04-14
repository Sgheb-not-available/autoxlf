"""
Estrai Risorse Umane da file Blumatica DVR (.blps) -> Excel (.xlsx)
Uso: python estrai_risorse_umane.py [cartella_con_blps]
     Se non si specifica la cartella, usa la directory corrente.
"""

import zipfile
import xml.etree.ElementTree as ET
import os
import sys
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Campi da estrarre per ogni risorsa umana
# Mappa: (etichetta colonna Excel) -> (possibili tag XML nel .blps)
# Blumatica usa nomi come "Nome", "Cognome", "CodiceFiscale", ecc.
CAMPI = {
    'Cognome': ['Cognome'],
    'Nome': ['Nome'],
    'ComuneNascita': ['ComuneNascita'],
    'DataNascita':  ['DataNascita'],
    'Sesso': ['Sesso'],
    'CF': ['CF'],
    'ComuneResidenza': ['ComuneResidenza'],
    'PR': ['PR'],
    'CAP': ['CAP'],
    'Indirizzo': ['Indirizzo'],
    'DataAssunzione': ['DataAssunzione'],
    'Note': ['Note'],
    'Immagine': ['Immagine'],
    'Mansioni': ['Mansioni'],
    'Email': ['Email']
}

HEADER_FISSI = list(CAMPI.keys())


def trova_testo(element, tag_list):
    """Cerca il testo di un tag tra diverse varianti possibili."""
    for tag in tag_list:
        found = element.find(f".//{tag}")
        if found is not None and found.text:
            return found.text.strip()
    return ""


def estrai_da_blps(percorso_blps):
    """
    Apre un .blps (ZIP), trova l'XML principale e restituisce:
    - nome_azienda (str)
    - lista di dict con i dati di ogni risorsa umana
    """
    nome_azienda = os.path.splitext(os.path.basename(percorso_blps))[0]
    risorse = []

    try:
        with zipfile.ZipFile(percorso_blps, 'r') as z:
            # Lista tutti i file nell'archivio
            nomi_file = z.namelist()
            print(nomi_file)

            # Cerca il file XML principale (spesso è il primo, o si chiama
            # "data.xml", "project.xml", o ha estensione .xml/.blpx)
            xml_candidates = [f for f in nomi_file if f.endswith(('.xml', '.blpx', '.blp'))]
            if not xml_candidates:
                xml_candidates = nomi_file  # prova tutti

            xml_principale = None
            for candidato in xml_candidates:
                try:
                    with z.open(candidato) as f:
                        contenuto = f.read()
                        # Il file principale contiene tag risorse umane
                        if any(kw in contenuto for kw in
                               [b'RisorseUmane', b'RisorsaUmana', b'Lavoratore',
                                b'Dipendente', b'Cognome', b'CodiceFiscale']):
                            xml_principale = candidato
                            xml_contenuto = contenuto
                            break
                except Exception:
                    continue

            if xml_principale is None:
                # Fallback: prende il file più grande
                sizes = {f: z.getinfo(f).file_size for f in nomi_file}
                xml_principale = max(sizes, key=sizes.get)
                with z.open(xml_principale) as f:
                    xml_contenuto = f.read()

            # Parse XML
            try:
                root = ET.fromstring(xml_contenuto)
            except ET.ParseError:
                # Prova a rimuovere BOM se presente
                xml_contenuto = xml_contenuto.lstrip(b'\xef\xbb\xbf')
                root = ET.fromstring(xml_contenuto)

            # Estrai nome azienda dal XML se disponibile
            for tag_az in ['NomeAzienda', 'RagioneSociale', 'Azienda', 'nomeAzienda']:
                el = root.find(f'.//{tag_az}')
                if el is not None and el.text:
                    nome_azienda = el.text.strip()
                    break

            # Trova il contenitore delle risorse umane
            container_tags = [
                'RisorseUmane', 'Lavoratori', 'Dipendenti', 'Personale',
                'risorseUmane', 'lavoratori'
            ]
            container = None
            for tag in container_tags:
                container = root.find(f'.//{tag}')
                if container is not None:
                    break

            # Tag per ogni singola risorsa
            item_tags = [
                'RisorsaUmana', 'Lavoratore', 'Dipendente', 'Persona',
                'risorsaUmana', 'lavoratore', 'Item', 'Record'
            ]

            elementi = []
            if container is not None:
                for tag in item_tags:
                    elementi = container.findall(tag)
                    if elementi:
                        break
                if not elementi:
                    elementi = list(container)  # tutti i figli diretti

            # Fallback: cerca nel documento intero
            if not elementi:
                for tag in item_tags:
                    elementi = root.findall(f'.//{tag}')
                    if elementi:
                        break

            for el in elementi:
                risorsa = {"Azienda": nome_azienda,
                           "File sorgente": os.path.basename(percorso_blps)}
                for colonna, tag_list in CAMPI.items():
                    risorsa[colonna] = trova_testo(el, tag_list)
                # Includi solo se ha almeno cognome o nome
                if risorsa.get("Cognome") or risorsa.get("Nome"):
                    risorse.append(risorsa)

    except zipfile.BadZipFile:
        print(f"'{percorso_blps}' non è un file ZIP valido — saltato.")
    except Exception as e:
        print(f"Errore su '{percorso_blps}': {e}")

    return nome_azienda, risorse


def crea_excel(tutte_risorse, percorso_output):
    """Crea il file Excel con un foglio riepilogo + un foglio per azienda."""
    wb = Workbook()

    # Stili
    colore_header = "1F4E79"
    colore_altriga = "DEEAF1"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell_font = Font(name="Arial", size=10)
    header_fill = PatternFill("solid", start_color=colore_header)
    alt_fill = PatternFill("solid", start_color=colore_altriga)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    def scrivi_foglio(ws, righe, titolo):
        ws.title = titolo[:31]  # Excel max 31 caratteri per nome foglio
        ws.freeze_panes = "A2"

        # Intestazioni
        for col_idx, header in enumerate(HEADER_FISSI, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border_thin

        # Dati
        for row_idx, risorsa in enumerate(righe, 2):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            for col_idx, header in enumerate(HEADER_FISSI, 1):
                valore = risorsa.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=valore)
                cell.font = cell_font
                cell.fill = fill
                cell.alignment = left
                cell.border = border_thin

        # Larghezze colonne
        larghezze = {
            "Azienda": 30, "File sorgente": 25, "Cognome": 18, "Nome": 16,
            "Codice Fiscale": 18, "Data di Nascita": 15, "Sesso": 8,
            "Mansione": 22, "Reparto / Sede": 20, "Data Assunzione": 15,
            "Tipo Contratto": 16, "Formazione": 12, "Sorveglianza San.": 16,
            "Nomina": 18, "Note": 25,
        }
        for col_idx, header in enumerate(HEADER_FISSI, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = larghezze.get(header, 15)
        ws.row_dimensions[1].height = 20

    # Foglio RIEPILOGO con tutte le aziende
    ws_riepilogo = wb.active
    scrivi_foglio(ws_riepilogo, tutte_risorse, "RIEPILOGO")

    # Un foglio per ogni azienda
    aziende = {}
    for r in tutte_risorse:
        az = r.get("Azienda", "Sconosciuta")
        aziende.setdefault(az, []).append(r)

    for nome_az, righe in sorted(aziende.items()):
        # Pulisci il nome per usarlo come nome foglio
        nome_pulito = "".join(c for c in nome_az if c not in r'\/*?:[]\x00-\x1f')[:31] or "Azienda"
        ws = wb.create_sheet(title=nome_pulito)
        scrivi_foglio(ws, righe, nome_pulito)

    wb.save(percorso_output)

# Entry point
def main():
    cartella_IN = input('Inserisci la cartella di input per dvr: ')
    cartella_IN = os.path.abspath(cartella_IN)

    cartella_OUT = input('Inserisci il percorso di destinazione file: ')
    cartella_OUT = os.path.abspath(cartella_OUT)

    print(f"\n{'='*60}")
    print(f"  Blumatica DVR -> Excel  |  Estrazione Risorse Umane")
    print(f"{'='*60}")
    print(f"  Cartella Input: {cartella_IN}\n")
    print(f"  Cartella Output: {cartella_OUT}\n")

    # Trova tutti i .blps nella cartella (ricorsivo)
    pattern = os.path.join(cartella_IN, "**", "*.blps")
    file_blps = glob.glob(pattern, recursive=True)

    if not file_blps:
        print("  Nessun file .blps trovato nella cartella specificata.")
        return

    print(f"  Trovati {len(file_blps)} file .blps\n")

    tutte_risorse = []
    for i, percorso in enumerate(sorted(file_blps), 1):
        nome_file = os.path.basename(percorso)
        print(f"  [{i}/{len(file_blps)}] Elaborazione: {nome_file}")
        nome_az, risorse = estrai_da_blps(percorso)
        print(f"        -> Azienda: {nome_az} | Risorse trovate: {len(risorse)}")
        tutte_risorse.extend(risorse)

    if not tutte_risorse:
        print("\n Nessuna risorsa umana trovata nei file elaborati.")
        print("  Possibile causa: struttura XML diversa dal previsto.")
        print("  Contatta il supporto allegando un file .blps di esempio.")
        return

    output = os.path.join(cartella_OUT, "risorse_umane_export.xlsx")
    crea_excel(tutte_risorse, output)

    print(f"\n{'='*60}")
    print(f"  Completato!")
    print(f"  Totale risorse estratte: {len(tutte_risorse)}")
    print(f"  File salvato: {output}")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()
